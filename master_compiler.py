import pandas as pd
import os
import datetime
import re
import shutil
import json
import difflib
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows

class MasterCompiler:
    def __init__(self):
        self.template_file = "Tracker_Template.xlsx"
        self.master_file = "TotalBattle_Master.xlsx"
        
        # LOGS
        self.chest_log = "chest_log.csv"
        self.roster_log = "roster_log.csv"
        self.mapping_file = "chest_mappings.json"
        self.player_map_file = "player_mappings.json" 
        self.ignore_file = "ignored_players.json" # NEW: Ignore list
        
        # MEMORY
        self.col_map = {} 
        self.point_map = {}
        self.prev_week_col_start = None 
        
        self.user_mappings = self.load_json(self.mapping_file)
        self.player_mappings = self.load_json(self.player_map_file)
        self.ignored_players = self.load_json(self.ignore_file) # NEW
        
        self.roster_names = self.load_roster_names()
        self.load_template_map()

    # --- SETUP & HELPERS ---
    def load_json(self, filepath):
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r') as f: 
                    data = json.load(f)
                    # Handle list format for ignore file
                    if isinstance(data, list): return set(data)
                    return data
            except: pass
        return {} if "ignore" not in filepath else set()

    def load_roster_names(self):
        names = set()
        if os.path.exists(self.roster_log):
            try:
                df = pd.read_csv(self.roster_log)
                if 'Player Name' in df.columns:
                    for n in df['Player Name'].dropna().unique():
                        # Apply Player Mapping immediately
                        raw = str(n).strip()
                        if raw in self.ignored_players: continue # NEW: Skip ignored
                        
                        clean = self.player_mappings.get(raw, raw)
                        if clean in self.ignored_players: continue # NEW: Skip ignored alias
                        
                        names.add(clean)
            except: pass
        return sorted(list(names))

    def get_clean_name(self, raw_name):
        if raw_name in self.ignored_players: return None # NEW

        # 1. Check Player Mappings (Aliases)
        if raw_name in self.player_mappings:
            mapped = self.player_mappings[raw_name]
            if mapped in self.ignored_players: return None
            return mapped
            
        # 2. Exact Match in known roster
        if raw_name in self.roster_names: return raw_name
        
        # 3. Fuzzy Match
        matches = difflib.get_close_matches(raw_name, self.roster_names, n=1, cutoff=0.90)
        if matches: return matches[0]
        
        return raw_name

    def load_template_map(self):
        if not os.path.exists(self.template_file): return
        try:
            wb = load_workbook(self.template_file, data_only=True)
            ws = wb.active 
            current_group = "Unknown"
            
            for col_idx in range(1, ws.max_column + 1):
                group_val = ws.cell(row=1, column=col_idx).value
                points_val = ws.cell(row=2, column=col_idx).value
                level_val = ws.cell(row=3, column=col_idx).value
                
                if group_val: 
                    current_group = str(group_val).strip()
                    if "Previous Week" in current_group: self.prev_week_col_start = col_idx
                
                if level_val:
                    l_str = str(level_val).lower().replace("lvl", "").strip()
                    key = f"{current_group} {l_str}".lower()
                    self.col_map[key] = col_idx
                    try:
                        p_str = str(points_val).lower().replace("pts", "").strip()
                        self.point_map[key] = int(p_str)
                    except: self.point_map[key] = 0
        except: pass

    def identify_chest(self, chest_name, chest_source):
        c_name = chest_name.strip()
        c_source = chest_source.strip()
        lookup_key = f"{c_name} | {c_source}" if "bank" in c_source.lower() else c_source
        if lookup_key in self.user_mappings: return self.user_mappings[lookup_key]['key']

        text = f"{c_name} {c_source}".lower()
        group = "other"
        if "common" in text or "barbarian" in text: group = "common crypt"
        elif "rare" in text: group = "rare crypt"
        elif "epic" in text: group = "epic crypt"
        elif "citadel" in text: group = "citadel"
        elif "heroic" in text: group = "heroic"
        elif "wooden" in text or "gold" in text: group = "bank chest"
        elif "tartaros" in text: group = "tartaros"
        
        level = ""
        match = re.search(r'(?:level|lvl)?[\s.]*(\d+)\b', text)
        if match:
            val = int(match.group(1))
            level = str(val)
            if group == "heroic":
                if 16 <= val <= 25: level = "16 - 25"
                elif 26 <= val <= 30: level = "26 - 30"
                elif val >= 31: level = "31+"
        return f"{group} {level}".strip().lower()

    def get_game_week_ranges(self):
        now_utc = datetime.datetime.utcnow()
        days_since_sun = (now_utc.weekday() + 1) % 7
        last_sun_date = now_utc.date() - datetime.timedelta(days=days_since_sun)
        this_week_reset = datetime.datetime.combine(last_sun_date, datetime.time(17, 0))
        
        if now_utc < this_week_reset:
            cw_start = this_week_reset - datetime.timedelta(days=7)
        else:
            cw_start = this_week_reset
            
        pw_start = cw_start - datetime.timedelta(days=7)
        pw_end = cw_start
        return cw_start, pw_start, pw_end

    def update_tracker(self, wb, min_score):
        ws = wb.active
        ws.title = "Tracker"
        cw_start, pw_start, pw_end = self.get_game_week_ranges()
        
        players = {}
        for r_name in self.roster_names:
            players[r_name] = {"cw_score": 0, "cw_chests": 0, "cw_counts": {}, "pw_score": 0, "pw_chests": 0}

        if os.path.exists(self.chest_log):
            try:
                df = pd.read_csv(self.chest_log)
                if not df.empty:
                    df['dt'] = pd.to_datetime(df['Timestamp'], dayfirst=True, errors='coerce')
                    
                    for _, row in df.iterrows():
                        if pd.isna(row['dt']): continue
                        row_dt = row['dt'].to_pydatetime()
                        if row_dt.tzinfo is None:
                            cw_start_naive = cw_start.replace(tzinfo=None)
                            pw_start_naive = pw_start.replace(tzinfo=None)
                            pw_end_naive = pw_end.replace(tzinfo=None)
                        else:
                            cw_start_naive = cw_start
                            pw_start_naive = pw_start
                            pw_end_naive = pw_end

                        raw_name = str(row['Player Name']).strip()
                        p_name = self.get_clean_name(raw_name)
                        
                        if not p_name: continue # Skip ignored / unmatched

                        if p_name not in players:
                             players[p_name] = {"cw_score": 0, "cw_chests": 0, "cw_counts": {}, "pw_score": 0, "pw_chests": 0}

                        c_name = str(row['Chest Name'])
                        c_src = str(row['Chest Source'])
                        key = self.identify_chest(c_name, c_src)
                        
                        if key in self.col_map:
                            pts = self.point_map.get(key, 0)
                            if row_dt >= cw_start_naive:
                                players[p_name]["cw_score"] += pts
                                players[p_name]["cw_chests"] += 1
                                players[p_name]["cw_counts"][key] = players[p_name]["cw_counts"].get(key, 0) + 1
                            elif pw_start_naive <= row_dt < pw_end_naive:
                                players[p_name]["pw_score"] += pts
                                players[p_name]["pw_chests"] += 1
            except: pass

        # ... (Rendering logic remains same) ...
        now_utc = datetime.datetime.utcnow()
        reset_today = now_utc.replace(hour=17, minute=0, second=0, microsecond=0)
        if now_utc < reset_today: last_reset = reset_today - datetime.timedelta(days=1)
        else: last_reset = reset_today
        diff = now_utc - last_reset
        hours = int(diff.total_seconds() // 3600)
        minutes = int((diff.total_seconds() % 3600) // 60)
        
        ws['A1'] = f"Last Updated\nDate: {now_utc.strftime('%d/%m/%Y')}\nReset: +{hours:02d}:{minutes:02d}"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 50 

        for cell_id in ['B1', 'C1', 'D1', 'E1']:
            if ws[cell_id].value: ws[cell_id].font = Font(size=16, bold=True)

        ws.column_dimensions['B'].width = 11
        ws.column_dimensions['C'].width = 11
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 11

        sorted_players = sorted(players.items(), key=lambda x: x[0].lower())
        start_row = 4
        center_align = Alignment(horizontal='center')
        bold_font = Font(bold=True)
        
        for i, (player, data) in enumerate(sorted_players):
            r = start_row + i
            c = ws.cell(row=r, column=1, value=player)
            c.font = bold_font
            ws.cell(row=r, column=2, value=data['cw_chests']).alignment = center_align
            c_score = ws.cell(row=r, column=3, value=data['cw_score'])
            c_score.font = bold_font
            c_score.alignment = center_align
            
            c_pct = ws.cell(row=r, column=4, value=f"=C{r}/E{r}")
            c_pct.number_format = '0%'
            c_pct.alignment = center_align
            ws.cell(row=r, column=5, value=min_score).alignment = center_align
            
            for key, col_idx in self.col_map.items():
                count = data["cw_counts"].get(key, 0)
                ws.cell(row=r, column=col_idx, value=count).alignment = center_align

            if self.prev_week_col_start:
                p_start = self.prev_week_col_start
                ws.cell(row=r, column=p_start, value=data['pw_chests']).alignment = center_align
                ws.cell(row=r, column=p_start+1, value=data['pw_score']).alignment = center_align
                if min_score > 0: prev_pct = data['pw_score'] / min_score
                else: prev_pct = 0
                c_prev_pct = ws.cell(row=r, column=p_start+2, value=prev_pct)
                c_prev_pct.number_format = '0%'
                c_prev_pct.alignment = center_align

        rule = ColorScaleRule(start_type='num', start_value=0, start_color='FF0000',
                              mid_type='num', mid_value=0.5, mid_color='FFFF00',
                              end_type='num', end_value=1, end_color='00FF00')
        ws.conditional_formatting.add(f"D4:D{start_row + len(sorted_players)}", rule)
        ws.freeze_panes = 'D4'
        ws.sheet_view.topLeftCell = 'A1'

    def update_growth(self, wb):
        if "Growth" in wb.sheetnames: del wb["Growth"]
        ws = wb.create_sheet("Growth")
        
        if not os.path.exists(self.roster_log): return

        try:
            df = pd.read_csv(self.roster_log)
            df['dt'] = pd.to_datetime(df['Scan Date'], dayfirst=True, errors='coerce')
            
            def clean_might(val):
                s = str(val).lower().replace(',', '')
                if 'k' in s: return float(re.sub(r'[^\d.]', '', s)) * 1000
                if 'm' in s: return float(re.sub(r'[^\d.]', '', s)) * 1000000
                try: return float(re.sub(r'[^\d.]', '', s))
                except: return 0

            df['MightVal'] = df['Might'].apply(clean_might)
            
            # --- NEW: Filter Ignored Players ---
            def resolve_name(n):
                return self.get_clean_name(str(n).strip())
            
            df['CleanName'] = df['Player Name'].apply(resolve_name)
            df = df.dropna(subset=['CleanName']) # Remove ignored/None
            # -----------------------------------

            cw_start, _, _ = self.get_game_week_ranges()
            if df['dt'].iloc[0].tzinfo is None: cw_start = cw_start.replace(tzinfo=None)
            
            df = df[df['dt'] >= cw_start]
            if df.empty: return

            headers = ["Player Name", "Start Date", "Start Might", "Current Date", "Current Might", "Growth", "Growth %"]
            ws.append(headers)
            # ... (Rest of Growth rendering is standard) ...
            for col in range(1, 8):
                ws.cell(row=1, column=col).font = Font(bold=True, size=12)
                ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

            stats = []
            for player, group in df.groupby('CleanName'):
                group = group.sort_values('dt')
                start_row = group.iloc[0]
                curr_row = group.iloc[-1]
                
                start_might = start_row['MightVal']
                curr_might = curr_row['MightVal']
                growth = curr_might - start_might
                pct = (growth / start_might) if start_might > 0 else 0
                
                stats.append([player, start_row['dt'].strftime("%d/%m"), start_might,
                              curr_row['dt'].strftime("%d/%m"), curr_might, growth, pct])
                
            stats.sort(key=lambda x: x[5], reverse=True)
            for row_data in stats: ws.append(row_data)
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 25
            
            for row in ws.iter_rows(min_row=2):
                for c_idx in [3, 5, 6]: row[c_idx-1].number_format = '#,##0'
                row[6].number_format = '0.00%'
                if row[5].value > 0: row[5].font = Font(color="009900")
                elif row[5].value < 0: row[5].font = Font(color="FF0000")

        except Exception as e: print(f"Growth Error: {e}")

    def update_events(self, wb, event_thresholds):
        files = glob.glob("event_*.csv")
        for f in files:
            try:
                # ... standard file prep ...
                raw_name = f.replace("event_", "").replace(".csv", "").replace("_", " ")
                clean_fname = raw_name.replace("   ", " - ") 
                target_score = event_thresholds.get(clean_fname, 0)
                tab_name = clean_fname[:30] 
                if tab_name in wb.sheetnames: del wb[tab_name]
                ws = wb.create_sheet(tab_name)
                
                df = pd.read_csv(f)
                def clean_score(s):
                    return int(re.sub(r'[^\d]', '', str(s))) if re.sub(r'[^\d]', '', str(s)) else 0
                
                df['ScoreVal'] = df['Score'].apply(clean_score)
                
                # --- NEW: Filter Ignored ---
                df['CleanName'] = df['Player Name'].apply(lambda x: self.get_clean_name(str(x).strip()))
                df = df.dropna(subset=['CleanName'])
                # ---------------------------
                
                final = df.groupby('CleanName')['ScoreVal'].max().reset_index()
                final = final.sort_values('ScoreVal', ascending=False)
                
                ws.append(["Rank", "Player Name", "Total Score", "Min Req"])
                # ... (Standard Event rendering) ...
                row_count = 0
                for idx, row in enumerate(dataframe_to_rows(final, index=False, header=False), 1):
                    ws.append([idx] + row + [target_score])
                    ws.cell(row=idx+1, column=3).number_format = '#,##0'
                    row_count = idx
                
                if target_score > 0:
                    rule = ColorScaleRule(start_type='num', start_value=0, start_color='FF0000', mid_type='num', mid_value=target_score/2, mid_color='FFFF00', end_type='num', end_value=target_score, end_color='00FF00')
                    ws.conditional_formatting.add(f"C2:C{row_count+1}", rule)
            except Exception as e: print(f"Event Error {f}: {e}")

    def run_update(self, min_score=1000, event_thresholds={}):
        try: shutil.copy(self.template_file, self.master_file)
        except: pass 
        try:
            wb = load_workbook(self.master_file)
            print("   -> Updating Tracker...")
            self.update_tracker(wb, min_score)
            print("   -> Updating Growth...")
            self.update_growth(wb)
            print("   -> Updating Events...")
            self.update_events(wb, event_thresholds)
            wb.save(self.master_file)
            print("✅ Master File Updated Successfully!")
            return True
        except Exception as e:
            print(f"❌ Critical Error: {e}")
            return False

if __name__ == "__main__":
    mc = MasterCompiler()
    mc.run_update()